from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

def create_excel_file(data):
    try:
        wb = load_workbook('data.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
    headers = ['Ім`я', 'Напої','Доїзд']
    for row_data in data:
        row = [row_data['name'], row_data['presence']]
        drinks = ['Вино червоне', 'Вино біле', 'Пиво', 'Горілка', 'Самогонка', 'Віскі', 'Ром', 'Не вживаю']
        row.append(', '.join(row_data['drinks']))
        row.append(', '.join(row_data['car']))
        sheet.append(row)
    for col_num, headers in enumerate(headers, start=1):
        col_letter = chr(65 + col_num)
        sheet[f'{col_letter}1'].font = Font(bold=True)
        sheet.column_dimensions[col_letter].width = 15

    

    wb.save('data.xlsx')

   
    print('Файл "data.xlsx" створено/оновлено успішно.')
def create_sorted_excel_file(data):
    wb = Workbook()
    sheet = wb.active
    headers = ['Ім`я', 'Присутність', 'Вино червоне', 'Вино біле', 'Пиво', 'Горілка', 'Самогонка', 'Віскі', 'Ром', 'Не вживаю','Доїзд']
    sheet.append(headers)
    for row_data in sorted(data, key=lambda x: x['name']):
        row = [row_data['name'], row_data['presence']]

        drinks = ['Вино червоне', 'Вино біле', 'Пиво', 'Горілка', 'Самогонка', 'Віскі', 'Ром', 'Не вживаю']
      
        for drink in drinks:
            if drink in row_data['drinks']:
                row.append('✔')
            else:
                row.append('✗')
        row.append(', '.join(row_data['car']))
        sheet.append(row)

    for col_num, headers in enumerate(headers, start=1):
        col_letter = chr(65 + col_num)
        sheet[f'{col_letter}1'].font = Font(bold=True)
        sheet.column_dimensions[col_letter].width = 15

    wb.save('sorted_data.xlsx')
    print('Файл "sorted_data.xlsx" створено/оновлено успішно.')
def create_calculator_excel_file(data):
    wb = Workbook()
    sheet = wb.active
    headers = ['Ім`я', 'Присутність', 'Вино червоне', 'Вино біле', 'Пиво', 'Горілка', 'Самогонка', 'Віскі', 'Ром', 'Не вживаю','На машині', 'Без машини']
    sheet.append(headers)
    for row_data in sorted(data, key=lambda x: x['name']):
        row = [row_data['name'], row_data['presence']]

        drinks = ['Вино червоне', 'Вино біле', 'Пиво', 'Горілка', 'Самогонка', 'Віскі', 'Ром', 'Не вживаю']
        cars=['Приїду на машині', 'Без машини']
        for drink in drinks:
            if drink in row_data['drinks']:
                row.append('1')
            else:
                row.append('0')
        for car in cars:
            if car in row_data['car']:
                row.append('1')
            else:
                row.append('0')
        sheet.append(row)

    for col_num, headers in enumerate(headers, start=1):
        col_letter = chr(65 + col_num)
        sheet[f'{col_letter}1'].font = Font(bold=True)
        sheet.column_dimensions[col_letter].width = 15
    global total_names, total_wines_red, total_wines_white, total_beers, total_rom, total_viski, total_no, total_vodka, total_samogon
    total_names = len(data)  # Загальна кількість учасників
    total_presence = sum(1 for row_data in data if row_data['presence'] == 'Присутній')
    total_absence = sum(1 for row_data in data if  row_data['presence'] == 'Відсутній')
    total_50 = sum(1 for row_data in data if row_data['presence'] == 'Пізніше')
    total_wines_red = sum(1 for row_data in data if 'Вино червоне' in row_data['drinks'])
    total_wines_white = sum(1 for row_data in data if 'Вино біле' in row_data['drinks'])
    total_beers = sum(1 for row_data in data if 'Пиво' in row_data['drinks'])
    total_vodka = sum(1 for row_data in data if 'Горілка' in row_data['drinks'])
    total_samogon = sum(1 for row_data in data if 'Самогонка' in row_data['drinks'])
    total_viski = sum(1 for row_data in data if 'Віскі' in row_data['drinks'])
    total_rom = sum(1 for row_data in data if 'Ром' in row_data['drinks'])
    total_no = sum(1 for row_data in data if 'Не вживаю' in row_data['drinks'])
    total_cars = sum(1 for row_data in data if 'Приїду на машині' in row_data['car'])
    total_no_cars = sum(1 for row_data in data if 'Без машини' in row_data['car'])
    
    total_presence = 0
    total_absence = 0
    total_50 = 0
    for row_data in data:
        if row_data['presence'] == 'Присутній':
            total_presence+=1
           
        elif row_data['presence'] == 'Відсутній':
            total_absence+=1
           
        elif row_data['presence'] == 'Пізніше':
            total_50+=1
        selected_total=total_50+total_presence+total_absence
            

    print(selected_total)
    print(type(total_wines_red))
    total_row = ['Підсумок',selected_total,total_wines_red, total_wines_white, total_beers, total_vodka, total_samogon, total_viski, total_rom, total_no, total_cars,total_no_cars]
    sheet.append(total_row)
    total_row_labels = ['Будуть', 'Не будуть', 'Не знаю', 'Відмітилось']
    total_row_values = [ total_presence, total_absence,total_50, selected_total]
    sheet.append(total_row_labels)
    sheet.append(total_row_values)
   


    wb.save('calculator_data.xlsx')
    print('Файл "calculator_data.xlsx" створено/оновлено успішно.')
#Видалення даних
def clear_excel_data():
    try:
        wb = load_workbook('data.xlsx')
        sheet = wb.active
        sheet.delete_rows(2, sheet.max_row)  # Видалити рядки від другого рядка до останнього
        wb.save('data.xlsx')
        print('Дані успішно видалено з таблиці.')
    except FileNotFoundError:
        print('Файл "data.xlsx" не знайдено.')
